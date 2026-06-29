from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from artifacts.artifact_manager import ensure_export_dir

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="2563EB")


def _coerce_sheet_spec(item: Any, index: int, fallback_title: str = "") -> dict[str, Any]:
    if isinstance(item, dict):
        return {
            "title": str(item.get("title") or item.get("name") or fallback_title or f"Sheet{index + 1}"),
            "headers": [str(h) for h in (item.get("headers") or [])],
            "rows": [list(row) for row in (item.get("rows") or [])],
        }
    if isinstance(item, (list, tuple)) and len(item) >= 3:
        return {
            "title": str(item[0] or f"Sheet{index + 1}"),
            "headers": [str(h) for h in item[1]],
            "rows": [list(row) for row in item[2]],
        }
    return {
        "title": fallback_title or f"Sheet{index + 1}",
        "headers": [],
        "rows": [],
    }


def normalize_excel_sheet_specs(
    title: str = "",
    headers: list[str] | None = None,
    rows: list[list] | None = None,
    sheets: list[Any] | None = None,
) -> list[dict[str, Any]]:
    if sheets:
        specs = [_coerce_sheet_spec(item, idx, title) for idx, item in enumerate(sheets)]
        if specs:
            return specs
    return [{
        "title": title or "Sheet1",
        "headers": [str(h) for h in (headers or [])],
        "rows": [list(row) for row in (rows or [])],
    }]


def excel_input_has_data(
    title: str = "",
    headers: list[str] | None = None,
    rows: list[list] | None = None,
    sheets: list[Any] | None = None,
) -> bool:
    for spec in normalize_excel_sheet_specs(title, headers, rows, sheets):
        if spec.get("headers") or spec.get("rows"):
            return True
    return False


def validate_excel_input(
    title: str = "",
    headers: list[str] | None = None,
    rows: list[list] | None = None,
    sheets: list[Any] | None = None,
) -> None:
    if excel_input_has_data(title, headers, rows, sheets):
        return
    raise ValueError(
        "Excel 缺少有效数据：请在 headers/rows 或 sheets 中提供完整表头与数据行。"
        "需要模型撰写分析内容时，请使用 agent.execute / office_excel_create 并填入完整表格数据。"
    )


def _write_sheet(ws: Worksheet, spec: dict[str, Any]) -> None:
    ws.title = (spec.get("title") or "Sheet1")[:31] or "Sheet1"
    headers = spec.get("headers") or []
    if headers:
        ws.append(headers)
        for cell in ws[1]:
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
    for row in spec.get("rows") or []:
        ws.append(list(row))
    for column in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 12), 42)


def create_excel_workbook(
    title: str = "",
    headers: list[str] | None = None,
    rows: list[list] | None = None,
    output_name: str = "workbook.xlsx",
    *,
    sheets: list[Any] | None = None,
) -> Path:
    out_dir = ensure_export_dir()
    path = out_dir / output_name
    specs = normalize_excel_sheet_specs(title, headers, rows, sheets)
    validate_excel_input(title, headers, rows, sheets)

    wb = Workbook()
    for idx, spec in enumerate(specs):
        ws = wb.active if idx == 0 else wb.create_sheet()
        _write_sheet(ws, spec)

    wb.save(path)
    return path
